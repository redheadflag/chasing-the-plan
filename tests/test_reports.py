from __future__ import annotations

import io
from urllib.parse import unquote

import openpyxl
from fastapi.testclient import TestClient

from tests.test_api import _workout_payload

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _xlsx_text(content: bytes) -> str:
    wb = openpyxl.load_workbook(io.BytesIO(content))
    return "\n".join(
        str(c) for row in wb.active.iter_rows(values_only=True) for c in row if c is not None
    )


def _seed_plan(client: TestClient, cat: dict[str, int]) -> int:
    a = client.post("/api/athletes", json={"name": "Иван", "note": "заметка"}).json()
    client.post("/api/workouts", json=_workout_payload(a["id"], cat)).raise_for_status()
    return a["id"]


def test_xlsx_report(client: TestClient, sample_catalog):
    aid = _seed_plan(client, sample_catalog)
    r = client.get(f"/api/athletes/{aid}/plan.xlsx")
    assert r.status_code == 200
    assert r.headers["content-type"] == XLSX_MIME
    assert "attachment" in r.headers["content-disposition"]
    assert len(r.content) > 0

    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    ws = wb.active
    text = "\n".join(
        str(c) for row in ws.iter_rows(values_only=True) for c in row if c is not None
    )
    assert "Среда (Crossfit)" in text
    assert "TURKISH GETUP" in text
    assert "4кг х 5" in text
    assert "20сек" in text  # time-based stage
    assert "Делаешь суперсетом." in text

    # hyperlink present on an exercise cell
    links = [c.hyperlink.target for row in ws.iter_rows() for c in row if c.hyperlink]
    assert "https://ex/tg" in links


def test_pdf_report(client: TestClient, sample_catalog):
    aid = _seed_plan(client, sample_catalog)
    r = client.get(f"/api/athletes/{aid}/plan.pdf")
    assert r.status_code == 200
    assert r.headers["content-type"] == "application/pdf"
    assert "attachment" in r.headers["content-disposition"]
    assert r.content[:5] == b"%PDF-"
    # embedded hyperlink annotation
    assert b"https://ex/tg" in r.content


def test_report_day_filter(client: TestClient, sample_catalog):
    aid = _seed_plan(client, sample_catalog)
    # workout is on WED; filtering to MON should yield an (empty) but valid report
    r = client.get(f"/api/athletes/{aid}/plan.xlsx?days=MON")
    assert r.status_code == 200
    assert len(r.content) > 0


def test_report_omits_athlete_name(client: TestClient, sample_catalog):
    aid = _seed_plan(client, sample_catalog)  # athlete name is "Иван"
    r = client.get(f"/api/athletes/{aid}/plan.xlsx")
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    text = "\n".join(
        str(c) for row in wb.active.iter_rows(values_only=True) for c in row if c
    )
    assert "Иван" not in text
    # the day header is still present
    assert "Среда (Crossfit)" in text


def _seed_two_weeks(client: TestClient, cat: dict[str, int]) -> int:
    """Week 1: WED (Crossfit). Week 2: FRI (Верх) + MON (Низ). Returns athlete id."""
    a = client.post("/api/athletes", json={"name": "Иван", "note": None}).json()
    client.post("/api/workouts", json=_workout_payload(a["id"], cat)).raise_for_status()
    for day, name in (("FRI", "Верх"), ("MON", "Низ")):
        w = _workout_payload(a["id"], cat)
        w["week"] = 2
        w["day_of_week"] = day
        w["name"] = name
        client.post("/api/workouts", json=w).raise_for_status()
    return a["id"]


def test_report_week_filter(client: TestClient, sample_catalog):
    aid = _seed_two_weeks(client, sample_catalog)

    # Single-week download: only that week's days, filename names the week, and
    # (single week) the sheet stays flat — no "Неделя N" separators inside.
    r = client.get(f"/api/athletes/{aid}/plan.xlsx?week=2")
    assert r.status_code == 200
    assert "Неделя 2" in unquote(r.headers["content-disposition"])
    text = _xlsx_text(r.content)
    assert "Пятница (Верх)" in text and "Понедельник (Низ)" in text
    assert "Среда" not in text  # week 1 excluded
    assert "Неделя" not in text  # single week -> no in-sheet week separators


def test_report_combined_has_week_headers(client: TestClient, sample_catalog):
    aid = _seed_two_weeks(client, sample_catalog)

    # Whole-plan export spans >1 week, so it gets "Неделя N" separators.
    r = client.get(f"/api/athletes/{aid}/plan.xlsx")
    assert r.status_code == 200
    text = _xlsx_text(r.content)
    assert "Неделя 1" in text
    assert "Неделя 2" in text
    assert "Среда (Crossfit)" in text
    assert "Пятница (Верх)" in text


def test_report_week_pdf_downloads(client: TestClient, sample_catalog):
    aid = _seed_two_weeks(client, sample_catalog)
    r = client.get(f"/api/athletes/{aid}/plan.pdf?week=1")
    assert r.status_code == 200
    assert r.content[:5] == b"%PDF-"
    assert "Неделя 1" in unquote(r.headers["content-disposition"])


def test_report_single_workout_only(client: TestClient, sample_catalog):
    a = client.post("/api/athletes", json={"name": "Иван", "note": None}).json()
    wed = client.post("/api/workouts", json=_workout_payload(a["id"], sample_catalog)).json()
    fri = dict(_workout_payload(a["id"], sample_catalog))
    fri["name"] = "Верх"
    fri["day_of_week"] = "FRI"
    client.post("/api/workouts", json=fri).raise_for_status()

    r = client.get(f"/api/athletes/{a['id']}/plan.xlsx?workout_id={wed['id']}")
    assert r.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    text = "\n".join(
        str(c) for row in wb.active.iter_rows(values_only=True) for c in row if c
    )
    assert "Среда (Crossfit)" in text
    assert "Пятница" not in text  # the other workout is excluded
    # single-workout download names the file after the workout, not the athlete
    cd = r.headers["content-disposition"]
    assert "Crossfit" in cd and "Иван" not in cd
