from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.worksheet.worksheet import Worksheet

from app.reports.assembler import PlanDoc, PlanRow

# Palette sampled from the sample plan (ARGB, no leading '#').
RED = "FFF00000"
BLUE = "FF0070C0"
GREEN = "FF60A830"
BLACK = "FF000000"

ATHLETE_FONT = Font(name="Calibri", bold=True, size=12, color=BLACK)
ATHLETE_NOTE_FONT = Font(name="Calibri", italic=True, size=8, color="FF666666")
DAY_FONT = Font(name="Calibri", bold=True, size=11, color=RED)
MUSCLE_FONT = Font(name="Calibri", bold=True, size=9, color=BLACK)
EX_BLUE = Font(name="Calibri", bold=True, size=9, color=BLUE)
EX_GREEN = Font(name="Calibri", bold=True, size=9, color=GREEN)
STAGE_TOP_FONT = Font(name="Calibri", size=9, color=BLACK)
NOTE_FONT = Font(name="Calibri", size=9, color=BLACK)
NOTES_HEAD_FONT = Font(name="Calibri", bold=True, size=9, color=RED)
NOTES_BODY_FONT = Font(name="Calibri", size=9, color=BLACK)
LEGEND_FONT = Font(name="Calibri", size=9, color=GREEN)

WRAP_TOP = Alignment(wrap_text=True, vertical="top")
STAGE_ALIGN = Alignment(wrap_text=True, vertical="center", horizontal="center")
NOTE_ALIGN = Alignment(wrap_text=True, vertical="top", indent=1)

THIN_BLACK = Side(style="thin", color=BLACK)


def _group_bands(rows: list[PlanRow]) -> list[list[PlanRow]]:
    bands: list[list[PlanRow]] = []
    for row in rows:
        if bands and bands[-1][0].band_id == row.band_id:
            bands[-1].append(row)
        else:
            bands.append([row])
    return bands


def render_xlsx(doc: PlanDoc) -> bytes:
    wb = Workbook()
    ws: Worksheet = wb.active
    ws.title = (doc.athlete_name or "Plan")[:31]
    ws.sheet_view.showGridLines = False
    # Print/export on one page wide (landscape) so all stage + note columns fit.
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_margins = PageMargins(
        left=0.2, right=0.2, top=0.25, bottom=0.25, header=0, footer=0
    )

    n = doc.max_stages
    total_cols = 2 + n + 1
    note_col = total_cols
    last_letter = get_column_letter(total_cols)

    r = 1
    if doc.athlete_note:
        ws.cell(row=r, column=1, value=doc.athlete_note).font = ATHLETE_NOTE_FONT
        r += 1
        r += 1  # spacer

    for day in doc.days:
        ws.merge_cells(f"A{r}:{last_letter}{r}")
        ws.cell(row=r, column=1, value=day.title).font = DAY_FONT
        r += 1

        # Superset blocks get a black rule immediately before and after them.
        # Two adjacent supersets share the boundary rule (no doubled line).
        line_rows: set[int] = set()
        for band in _group_bands(day.rows):
            start_r = r
            for i, row in enumerate(band):
                is_first = i == 0
                mg = ws.cell(row=r, column=1, value=row.muscle_group)
                mg.font = MUSCLE_FONT
                mg.alignment = WRAP_TOP

                ex = ws.cell(row=r, column=2, value=row.exercise_name)
                ex.font = EX_GREEN if row.highlight else EX_BLUE
                ex.alignment = WRAP_TOP
                if row.exercise_url:
                    ex.hyperlink = row.exercise_url

                for s in range(n):
                    cell = ws.cell(row=r, column=3 + s)
                    cell.alignment = STAGE_ALIGN
                    if s < len(row.stages):
                        stage = row.stages[s]
                        cell.value = f"{stage.top}\n{stage.bottom}"
                        cell.font = STAGE_TOP_FONT

                note_text = row.note or ""
                if is_first and row.is_superset and row.block_note:
                    note_text = (
                        f"{note_text}\n{row.block_note}" if note_text else row.block_note
                    )
                nc = ws.cell(row=r, column=note_col, value=note_text or None)
                nc.font = NOTE_FONT
                nc.alignment = NOTE_ALIGN

                ws.row_dimensions[r].height = 30
                r += 1
            if band[0].is_superset:
                line_rows.add(start_r)  # rule before the superset
                line_rows.add(r)        # rule after the superset (top of next row)

        for rr in line_rows:
            for c in range(1, total_cols + 1):
                ws.cell(row=rr, column=c).border = Border(top=THIN_BLACK)

        if day.notes:
            r += 1
            ws.cell(row=r, column=1, value="Примечание:").font = NOTES_HEAD_FONT
            r += 1
            for line in day.notes.splitlines():
                if line.strip():
                    ws.cell(row=r, column=1, value=line).font = NOTES_BODY_FONT
                    r += 1
        r += 1  # spacer between days

    if any(row.highlight for day in doc.days for row in day.rows):
        ws.cell(
            row=r,
            column=1,
            value="Зелёным цветом выделены упражнения, которые нужно снять на видео.",
        ).font = LEGEND_FONT

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 34
    for s in range(n):
        ws.column_dimensions[get_column_letter(3 + s)].width = 12
    ws.column_dimensions[last_letter].width = 34

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
